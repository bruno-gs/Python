import openpyxl
import os

# Create a new blank excel workbook
wb = openpyxl.Workbook()

# To know the name of the sheets
wb.get_sheet_names()

# Return the spreadsheet
sheet = wb.get_sheet_by_name('Sheet')

# add data to a cell
sheet['A1'] = 42
sheet['A2'] = 'Hello'

os.chdir('pass the path')

# save the file with the modifications
wb.save('example.xlsx')

# add new sheets
sheet2 = wb.create_sheet()

# To change the name of the sheet
sheet2.title = 'My new Sheet Name'

# You can pass the index of the sheet and its name in the creat_sheet method
sheet3 = wb.create_sheet(index=0, title='My other Sheet')
    # in this case, I put the 'My other Sheet' at the begining, in the first position
